import pandas as pd
import requests
from io import BytesIO
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.dataframe import dataframe_to_rows

def limpiar_nombre(nombre):
    # Función para limpiar caracteres no deseados en el nombre del archivo de imagen
    caracteres_no_deseados = r'\/:*?"<>|'
    for caracter in caracteres_no_deseados:
        nombre = nombre.replace(caracter, '_')
    return nombre

def descargar_imagen(url, ruta_guardado):
    try:
        response = requests.get(url)
        response.raise_for_status()
        img = Image.open(BytesIO(response.content))
        img.save(ruta_guardado)
        return True
    except (requests.RequestException, IOError) as e:
        print(f'Error al descargar la imagen desde: {url}')
        print(str(e))
        return False
    except Exception as e:
        print(f'Error inesperado al procesar la URL: {url}')
        print(str(e))
        return False

def exportar_datos_con_imagenes():
    archivo_excel = '/Users/mimac/Downloads/TEST2/Script_Comparador_excel/tabla_destino_actualizada.xlsx'
    nombre_hoja = 'Sheet1'

    # Leer el archivo Excel
    df = pd.read_excel(archivo_excel, sheet_name=nombre_hoja)

    # Agregar columna nueva para las rutas de las imágenes descargadas
    df['RutaImagen'] = ''

    # Descargar las imágenes y guardar las rutas en el DataFrame
    for i, url_celda in enumerate(df['Main']):
        if pd.notnull(url_celda):  # Verificar si la URL no es nula
            urls = [url.strip() for url in url_celda.split(',')]  # Dividir las URLs separadas por comas en una lista
            rutas_imagenes = []

            for j, url in enumerate(urls, start=1):
                nombre_imagen = f"imagen_{df.at[i, 'SKU']}_{j}.jpg"  # Nombre del archivo de imagen basado en el valor de la columna SKU
                ruta_imagen = f'/Users/mimac/Downloads/TEST2/Scrip_comparador_excel{nombre_imagen}'  # Ruta completa de la imagen descargada

                if descargar_imagen(url, ruta_imagen):
                    rutas_imagenes.append(ruta_imagen)
                else:
                    rutas_imagenes.append('Error al descargar la imagen')

            df.at[i, 'RutaImagen'] = ','.join(rutas_imagenes)  # Asignar la ruta de la imagen al DataFrame

    # Crear un nuevo archivo Excel
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = nombre_hoja

    # Insertar los encabezados de las columnas en el archivo Excel
    encabezados = list(df.columns)
    for j, encabezado in enumerate(encabezados):
        cell = worksheet.cell(row=1, column=j+1)  # Primera fila, columna j+1
        cell.value = encabezado

    # Insertar los datos y las imágenes en el archivo Excel
    for i, fila in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):  # Comenzar desde la fila 2
        for j, valor in enumerate(fila, start=1):
            cell = worksheet.cell(row=i, column=j)
            cell.value = valor

        rutas_imagenes = df.at[i-2, 'RutaImagen'].split(',')  # Obtener las rutas de las imágenes correspondientes
        if any(pd.notnull(ruta_imagen) and ruta_imagen != 'Error al descargar la imagen' for ruta_imagen in rutas_imagenes):
            try:
                for idx, ruta_imagen in enumerate(rutas_imagenes, start=len(encabezados) + 1):
                    if pd.notnull(ruta_imagen) and ruta_imagen != 'Error al descargar la imagen':
                        img = Image.open(ruta_imagen)
                        img = img.resize((90, 90))  # Redimensionar la imagen a 90x90 píxeles
                        img_excel = ExcelImage(img)
                        img_excel.width = 90
                        img_excel.height = 90
                        cell = worksheet.cell(row=i, column=idx)  # Columna después de los datos
                        worksheet.add_image(img_excel, cell.coordinate)  # Insertar la imagen al lado de la URL
            except Exception as e:
                print(f'Error al insertar la imagen en el archivo Excel: {rutas_imagenes}')
                print(str(e))

    # Ajustar el ancho de las columnas
    for column in worksheet.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Guardar el archivo Excel
    nombre_archivo_nuevo = 'Vaporizadores_con_imagenesTEST1.xlsx'
    workbook.save(nombre_archivo_nuevo)
    print(f'Se ha guardado el archivo: {nombre_archivo_nuevo}')

# Llama a la función
exportar_datos_con_imagenes()
